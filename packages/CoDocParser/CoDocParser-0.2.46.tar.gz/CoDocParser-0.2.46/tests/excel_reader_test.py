# Press the green button in the gutter to run the script.
import os
import pytest
from openpyxl import load_workbook


class TestExcelReaderParser:
    """
    测试Excel文档解析器
    """
    def test_text_block_parse(self):
        """
        测试文本块解析
        """
        wb = load_workbook(os.getcwd() + '/tests/files/cma.xlsx')
       
        print(wb.sheetnames)
        a_sheet = wb['Table 1']
        rows = a_sheet.max_row
        cols = a_sheet.max_column

        dict={}
        # 因为按行，所以返回A1, B1, C1这样的顺序
        # for row in a_sheet.rows:
        #     for cell in row:
        #         for val in cell.value.spiltlines()
        #
        # # A1, A2, A3这样的顺序
        # for column in a_sheet.columns:
        #     for cell in column:
        #         print(cell.value)

        for r in range(1,rows):
           for c in range(1,cols):
                if a_sheet.cell(r,c).value is not None:
                    if  str(a_sheet.cell(r,c).value).find('\n')>0 :
                        for v in a_sheet.cell(r,c).value.splitlines():
                            # print(r,c,v)
                            if str(v).find(": ")>0:
                                dict[r,c,str(v).split(": ")[0]]=str(v).split(": ")[1]
                            else:
                                dict[r, c] = v
                    else:
                        if str(a_sheet.cell(r,c).value).find(": ") > 0:
                            dict[r, c, str(a_sheet.cell(r,c).value).split(": ")[0]] = str(a_sheet.cell(r,c).value).split(": ")[1]
                        else:
                            dict[r, c] = a_sheet.cell(r,c).value
                         # print(r, c, a_sheet.cell(r,c).value)
        assert wb is not None
        for key,value in dict.items():
            print('{key}:{value}'.format(key=key,value=value))


if __name__ == '__main__':
    pytest.main("-q --html=report.html")
