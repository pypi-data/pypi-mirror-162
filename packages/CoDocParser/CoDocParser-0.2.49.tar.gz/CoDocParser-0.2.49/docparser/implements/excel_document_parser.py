import openpyxl
import os
import re
from docparser.core.document_parser_base import DocumentParserBase
from docparser.config.excel.table_block_config import TableBlockConfigSchema


class ExcelDocumentParser(DocumentParserBase):
    """
    Excel文档解析器
    """

    def __init__(self, file, configs):
        """
        初始化
        :param file:文件路径
        :param configs: 配置
        """
        self._file = file
        self._configs = configs

        if not os.path.exists(file):
            raise FileNotFoundError

        work_book = openpyxl.load_workbook(file, read_only=True)
        self._sheet = work_book.worksheets[0]

    def parse(self):
        """
        根据配置抽取数据
        :return: 返回抽取的数据
        """

        data = {}
        errors = {}
        keylist = list(self._configs.keys())
        for key in self._configs.keys():
            keyconfig = self._configs[key]
            index = keylist.index(key)
            current_Keyword = keyconfig["keyword"]
            next_Keyword = ""

            if index < len(keylist) - 1:
                nextKey = keylist[index + 1]
                next_Keyword = self._configs[nextKey]["keyword"]

            result = self.search_keyword(keyconfig, next_Keyword)
            # if keyconfig.__contains__("isgroup")==False:
            data[key] = result[0], result[1], result[2]
            if keyconfig.__contains__("isgroup"):
                childkeys = list(keyconfig["child"].keys())
                if keyconfig.__contains__("child_mode"):
                    lines=result[0].split('#JOIN_AS_LINE#')
                    childgroup=[]
                    for line in lines:
                        lineTextArrs=line.replace(' ','#').replace('#JOIN_AS_COLUMN#','###').replace('\n','##')
                        kresults='@'.join('@'.join(lineTextArrs.split('###')).split('@@')).replace('@@','@').replace('#',' ').split('@')

                        child={}
                        for childkey in keyconfig["child"].keys():
                            childindex = childkeys.index(childkey)
                            if len(kresults) > childindex:
                                child[childkey] = kresults[childindex]
                                    #, result[0]
                            else:
                                child[childkey] = ''
                                    #, result[0]
                        childgroup.append(child)
                        #print(lineTextArrs.replace('@@','#').replace('####','@'))
                        #print(re.split(r"#@+", lineTextArrs.replace('@@','#')))
                    data[key]=childgroup,result[0]
                else:
                    results = result[0].splitlines()
                    if keyconfig.__contains__("spilt_mode"):
                       results = result[0].split(" ", len(childkeys) - 1)
                    for childkey in keyconfig["child"].keys():
                         childindex = childkeys.index(childkey)
                    # if childkey == 'billtype':
                    #     print(results)
                         if len(results) > childindex:
                                data[childkey] = results[childindex].strip(), result[0]
                         else:
                              data[childkey] = '', result[0]

        return data, errors

    def search_keyword(self, keyconfig, nextKeyword):
        keyword = keyconfig["keyword"]

        loc = self.loc_keyword(keyconfig, nextKeyword)
        arr = []
        r = loc[0][0]

        if keyconfig.__contains__("value_local_mode"):
            #print(loc, keyconfig)
            if keyconfig["value_local_mode"] == "next_rows":
                for r1 in range(loc[0][0] + 1, loc[1][0]):
                    arr1 = []
                    for cc in range(1, self._sheet.max_column):
                        cell1 = self._sheet.cell(r1, cc)
                        # print(r1,cc)
                        if cell1.value is not None:
                            arr1.append(str(cell1.value))
                            #print(cell1.row, cell1.column, cell1.column_letter,'$$$$', str(cell1.value))
                    arr.append('#JOIN_AS_COLUMN#'.join(arr1))
                return '#JOIN_AS_LINE#'.join(arr), '#JOIN_AS_LINE#'.join(arr), loc
        else:
            for c in range(loc[0][1], loc[1][1] + 1):
                cell = self._sheet.cell(r, c)
                if cell.value is not None:
                    arr.append(str(cell.value))
        mtext = ' '.join(arr)
        match = mtext

        if str(mtext).find(keyword) > -1:
            match = mtext[mtext.rfind(keyword) + len(keyword) + 1:]
        # if keyword.find("LOAD PICKUP POOL ADDRESS") > -1:
        #     print(keyword, nextKeyword, match.rfind(nextKeyword), mtext, match, len(arr))
        if (nextKeyword != "") & (match.find(nextKeyword) > -1):
            # print(keyword,nextKeyword,match.rfind(nextKeyword),mtext,match)
            match = match[:match.rfind(nextKeyword)]
        return match.lstrip().rstrip(), mtext, loc

    def loc_keyword(self, keyconfig, nextKeyword):
        keyword = keyconfig["keyword"]
        loc_row = 0
        loc_column = 0
        nextloc_row = self._sheet.max_row
        nextloc_column = self._sheet.max_column
        times = 0
        nextKeyowrd_times = 0

        if keyconfig.__contains__("stop_keyword"):
            nextKeyword=keyconfig["stop_keyword"]

        for row in self._sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    if str(cell.value).find(keyword) > -1:
                        if times == 0:
                            loc_row = cell.row
                            loc_column = cell.column
                            times = 1
                            continue
                    else:
                        if str(cell.value).find(nextKeyword) > -1:
                            if times > 0:
                                if nextKeyowrd_times == 0:
                                    nextloc_row = cell.row
                                    nextloc_column = cell.column
                                    nextKeyowrd_times = 1
                                    continue

        if keyconfig.__contains__("value_local_mode"):
            if keyconfig["value_local_mode"] != "next_rows":
                nextloc_row = loc_row
                nextloc_column = self._sheet.max_column
            else:
                nextloc_row = nextloc_row
        else:
            if loc_row != nextloc_row:
                nextloc_row = loc_row
                nextloc_column = self._sheet.max_column

        if loc_column > nextloc_column:
            nextloc_column = self._sheet.max_column
        # 预读取下个单元格的数据
        if loc_column == nextloc_column:
            nextloc_column = loc_column + 1

        # print(keyword,nextKeyword,[(loc_row,loc_column),(nextloc_row,nextloc_column)])
        return [(loc_row, loc_column), (nextloc_row, nextloc_column)]

    def _pre_check_and_process_table_config(key, table_config) -> object:
        """
        检查表格配置是否正确
         :param table_config:文本域配置
         :return: 返回文本域配置提取的内容
        """
        schema = TableBlockConfigSchema()
        return schema.validate(table_config)

    def _extract_text(self, text_config) -> object:
        """
        提取文本域数据
        :param text_config:文本域配置
        :return: 返回文本域配置提取的内容
        """
        # text_block_extractor = ExcelTextBlockExtractor(self._sheet, text_config)
        # errs = text_block_extractor.check()
        # if len(errs.keys()) > 0:
        #     return "", errs
        #
        # text = text_block_extractor.extract()
        # text = self._text_process(text, text_config)
        #
        # return text, None

    def _extract_table(self, table_config) -> object:
        """
        提取表格数据
        :param table_config: 表格配置
        :return: 返回配置提取表格的数据
        """

        # rect = table_config["rect"]
        # items = []
        # original_columns_maps = table_config["original_columns_maps"]
        # extract_columns = table_config["extract_columns"]
        # top = original_columns_maps[list(original_columns_maps.keys())[0]]["row"] + 1
        #
        # row_spans = rect["bottom"] - rect["top"]
        # for row in range(top, rect["top"] + row_spans):
        #     item = {}
        #     for col_name in extract_columns.keys():
        #         col_title = extract_columns[col_name]["title"]
        #         col_map = original_columns_maps[col_title]
        #         text = ''
        #         for col in range(col_map["col"], col_map["col"] + col_map["span"]):
        #             text = text + ExcelExtractorUtils.get_sheet_cell_value(self._sheet, row,
        #                                                                    original_columns_maps[col_title]["col"])
        #
        #         # 设置单元格值
        #         item[col_name] = self._cell_value_process(text, row, col, table_config["extract_columns"][col_name])
        #
        #     # 添加表格行数据
        #     items.append(item)
        #
        # return items

    def _text_process(self, text, text_config):
        """
        文本结果处理
        :param text:区域文本数据
        :param text_config: 配置
        :return:
        """
        if text is None or len(text) == 0:
            return ""

        return text

    def _cell_value_process(self, text, row, col, col_config):
        """
        文本结果处理
        :param text:区域文本数据
        :param row:行
        :param col:列
        :param col_config: 列配置
        :return:
        """
        if text is None or len(text) == 0:
            return ""

        return text


if __name__ == '__main__':
    cma_config = {
        # VESSEL:CMA CGM LYRA
        "vessel_dischport_placeofreceipt": {
            "isgroup": True,
            "child": {
                "vessel": {"keyword": "VESSEL"},
                "dischport": {"keyword": "OPERATIONAL DISCH. PORT"},
                "placeofreceipt": {"keyword": "PLACE OF RECEIPT"},
            },
            "keyword": "VESSEL:\nOPERATIONAL DISCH. PORT: PLACE OF RECEIPT:",
        },
        # "vessel": {"keyword": "VESSEL"},
        # VOYAGE:  0TX92W1MA
        "voyage": {"keyword": "VOYAGE"},
        # POD ETA:  07/15/2021
        "podeta": {"keyword": "POD ETA"},
        # OPERATIONAL DISCH.PORT: LOS ANGELES, CA
        # "dischport": {"keyword": "OPERATIONAL DISCH. PORT"},
        # PLACE OF RECEIPT:
        # "placeofreceipt": {"keyword": "PLACE OF RECEIPT"},
        # FPD ETA:
        "fpdeta": {"keyword": "FPD ETA"},
        # OPERATIONAL LOAD PORT: XIAMEN
        "operationalloadport": {
            "isgroup": True,
            "child": {
                "operationalloadport": {"keyword": "OPERATIONAL LOAD PORT"},
                # PLACE OF DELIVERY:
                "placeofdelivery": {"keyword": "PLACE OF DELIVERY"}
            },
            "keyword": "OPERATIONAL LOAD PORT:\nPLACE OF DELIVERY:",
            "value_local_mode": "next_column"
        },

        # DEST.CARG MODE: Port
        "dest.cargmode": {"keyword": "DEST.CARG MODE"},
        "destionation_itnumber": {
            "isgroup": True,
            "child": {
                # DESTINATION: LOS ANGELES, CA
                "destination": {"keyword": "DESTINATION"},
                # IT NUMBER: Local Clear
                "itnumber": {"keyword": "IT NUMBER"},
            },
            "keyword": "DESTINATION:\nIT NUMBER:",
        },
        # PLACE OF ISSUE: LOS ANGELES, CA
        "placeofissue": {"keyword": "PLACE OF ISSUE"},
        # IT ISSUED DATE:
        "itissueddate": {"keyword": "IT ISSUED DATE"},
        # LOAD PICKUP POOL ADDRESS: FENIX MARINE TERMINAL
        "loadpickuppooladdress": {"keyword": "LOAD PICKUP POOL ADDRESS"},
        # CLEARANCE POINT:   LOS ANGELES, CA
        "clearancepoint": {"keyword": "CLEARANCE POINT"},
        # FIRMS CODE:    Y257
        "firmscode": {"keyword": "FIRMS CODE"},
        # EMPTY RETURN DEPOT: Please Check   https://apps.usa.cma-cgm.com/econtainer/ daily
        "emptyreturndepot": {"keyword": "EMPTY RETURN DEPOT"},
        # RELEASE DATE:    07/23/2021
        "releasedate": {"keyword": "RELEASE DATE"},
        # PAYMENT RECEIVED:    NO
        "paymentreceived": {"keyword": "PAYMENT RECEIVED"},
        # OBL RECEIVED:YES
        "oblreceived": {"keyword": "OBL RECEIVED"},
        "SCAC_B/L#_BILL TYPE": {"isgroup": True,
                                "child": {
                                    # SCAC:CMDU
                                    "scac": {"keyword": "SCAC"},
                                    # B/L#:XIA0707492
                                    "billno": {"keyword": "B/L #"},
                                    # BILL TYPE:Waybill
                                    "billtype": {"keyword": "BILL TYPE"},

                                },
                                "keyword": "SCAC     B/L #                          BILL TYPE",
                                "spilt_mode": " "
                                },
        "hblno": {"keyword": "NVOCC/House bill info"},
        "containers": {
            "isgroup": True,
            "keyword": "CONTAINER  #        SEAL  #             SIZE/TYPE  PIECE QTY & TYPE    WEIGHT            MEASURE",
            "child": {
                # CONTAINER# : TRHU7122680
                "containerno": {"keyword": "CONTAINER#"},
                # SEAL#:  C0135997
                "sealno": {"keyword": "SEAL#"},
                # SIZE/TYPE: 40HC
                "size": {"keyword": "SIZE/TYPE"},
                # PIECE QTY&TYPE: 1060 CARTONS
                "qty": {"keyword": "PIECE QTY&TYPE"},
                # WEIGHT: 15909 LBS
                "weight": {"keyword": "WEIGHT"},
                # MEASURE:2398.00 FTQ
                "measure": {"keyword": "MEASURE"},
                # FREE BUSINESS DAYS AT PORT:
                "freebusinessdaysatport": {"keyword": "FREE BUSINESS", "replace_str": " DAYS AT PORT"},
                # LAST FREE DAY AT RAMP:
                "lastfreedayatramp": {"keyword": "LAST FREE ", "replace_str": " DAYS AT RAMP"},
                # PICKUP#:
                "pickupno": {"keyword": "PICKUP#"}
            },
            "value_local_mode": "next_rows",
            "child_mode": "multi"
        },
        "notes": {"keyword": "PLEASE NOTE :", "value_local_mode": "next_rows","stop_keyword":"SHIPPER" }

        # "containers": {
        #     # CONTAINER# : TRHU7122680
        #     "containerno": {"keyword": "CONTAINER#"},
        #     # SEAL#:  C0135997
        #     "sealno": {"keyword": "SEAL#"},
        #     # SIZE/TYPE: 40HC
        #     "size": {"keyword": "SIZE/TYPE"},
        #     # PIECE QTY&TYPE: 1060 CARTONS
        #     "qty": {"keyword": "PIECE QTY&TYPE"},
        #     # WEIGHT: 15909 LBS
        #     "weight": {"keyword": "WEIGHT"},
        #     # MEASURE:2398.00 FTQ
        #     "measure": {"keyword": "MEASURE"},
        #     # FREE BUSINESS DAYS AT PORT:
        #     "freebusinessdaysatport": {"keyword": "FREE BUSINESS DAYS AT PORT"},
        #     # LAST FREE DAY AT RAMP:
        #     "lastfreedayatramp": {"keyword": "LAST FREE DAY AT RAMP"},
        #     # PICKUP#:
        #     "pickupno": {"keyword": "PICKUP#"}
        # }
    }
    converter = ExcelDocumentParser(
        r"D:\Sources\gitlab\third-services\cots.libs\cots.libs.docparser\tests\files\cma.xlsx",
        cma_config)
    data = converter.parse()
    #print(data)
    ldata={}
    for item in data:
        for m in item.keys():
            #print('{key}:{value}'.format(key=m, value=item[m][0]))
            ldata[m]=item[m][0]

    print(ldata)

