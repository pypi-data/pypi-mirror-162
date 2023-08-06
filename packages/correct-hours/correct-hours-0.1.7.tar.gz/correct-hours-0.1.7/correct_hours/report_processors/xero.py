from openpyxl.styles import Font


class XeroReportProcessor:

    def __init__(self, workbook):
        self.workbook = workbook
        self.original_sheet = self.workbook.active
        self.new_sheet = self.workbook.copy_worksheet(from_worksheet=self.original_sheet)

    def add_new_column_headings(self):
        # new column headings
        self.new_sheet.cell(5, 15, "Old total").font = Font(bold=True)
        self.new_sheet.cell(5, 16, "New total").font = Font(bold=True)
        self.new_sheet.cell(5, 17, "Days worked").font = Font(bold=True)

    def add_up_hours(self, current_week_ending, row_start):
        rows_added = 0
        total_hours = 0
        for row in self.new_sheet.iter_rows(min_row=row_start, min_col=1, max_col=14, values_only=True):
            week_ending = row[0]
            if week_ending != current_week_ending:
                break
            total = row[13]
            total_hours += total
            rows_added += 1
        # print(f"Total hours for {current_week_ending} is {total_hours} within {rows_added} rows")
        return rows_added, total_hours

    def correct_hours(self, row_start, row_offset, overtime):
        # print("Hours need to be corrected")
        time_left = overtime
        for col_idx in range(13, 7, -1):
            for row_idx in range((row_start + row_offset) - 1, row_start - 1, -1):
                value = self.new_sheet.cell(row_idx, col_idx).value
                if value <= 0:
                    continue
                corrected_value = value - time_left
                if corrected_value < 0:
                    time_left = corrected_value * -1
                    corrected_value = 0
                    self.new_sheet.cell(row_idx, col_idx, corrected_value).font = Font(bold=True, color="FF0000")
                else:
                    self.new_sheet.cell(row_idx, col_idx, corrected_value).font = Font(bold=True, color="FF0000")
                    return

    def process_rows(self):
        row_count = 0
        for idx, row in enumerate(self.new_sheet.iter_rows(min_row=6, min_col=1, max_col=14, values_only=True)):
            if idx < row_count:
                continue
            week_ending = row[0]
            start_row_idx = idx + 6
            rows_added, total_hours = self.add_up_hours(week_ending, start_row_idx)
            end_row_idx = (start_row_idx + (rows_added - 1))
            # old total of hours
            self.new_sheet.cell(end_row_idx, 15, total_hours)
            did_overtime = False
            if total_hours > 38:
                overtime = total_hours - 38
                self.correct_hours(start_row_idx, rows_added, overtime)
                did_overtime = True
            row_count += rows_added
            # new total of hours
            new_total_of_hours = self.new_sheet.cell(end_row_idx, 16, f"=SUM(G{start_row_idx}:M{end_row_idx})")
            if did_overtime:
                new_total_of_hours.font = Font(bold=True, color="FF0000")
            # number of days worked
            self.new_sheet.cell(end_row_idx, 17, f"=P{end_row_idx}/7.6")

    def process(self):
        self.add_new_column_headings()
        self.process_rows()