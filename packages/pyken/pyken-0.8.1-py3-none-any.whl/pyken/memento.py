import numpy as np, pandas as pd

from sklearn.model_selection import train_test_split
from sklearn.tree import _tree, DecisionTreeClassifier

from carpe import *


####################################################################################################


class autoscorecard:


    def __init__(self, target_name='target', id_columns=[],
    autogrp_max_groups=5, autogrp_min_pct=0.05,
    flag_train_test=[], test_size=0.3, seed=123, stratify=True, stratify_var='',
    features=[], candidate_vars=[], excluded_vars=[], included_vars=[],
    selection_method='stepwise', selection_metric='pvalue',
    selection_threshold= 0.01, selection_max_iters=12,
    selection_muestra_test=True, selection_show='gini',
    user_breakpoints={}, log_mode=False, create_excel=True,
    save_whole_tables=False, save_all_autogroupings=False):

        self.target_name = target_name
        self.id_columns = id_columns

        self.autogrp_max_groups = autogrp_max_groups
        self.autogrp_min_pct = autogrp_min_pct

        self.flag_train_test = flag_train_test
        self.test_size = test_size
        self.seed = seed
        self.stratify = stratify
        self.stratify_var = stratify_var

        self.features = features
        self.candidate_vars = candidate_vars
        self.excluded_vars = id_columns + excluded_vars
        self.included_vars = included_vars

        self.selection_method = selection_method
        self.selection_metric = selection_metric
        self.selection_threshold = selection_threshold
        self.selection_max_iters = selection_max_iters
        self.selection_muestra_test = selection_muestra_test
        self.selection_show = selection_show

        self.user_breakpoints = user_breakpoints
        self.log_mode = log_mode
        self.create_excel = create_excel
        self.save_whole_tables = save_whole_tables
        self.save_all_autogroupings = save_all_autogroupings


    def fit(self, X, y):
        
        N = 150

        if self.log_mode: pre_text = '123: '
        else: pre_text = ''

        if self.flag_train_test != []:

            try: a, b, c = self.flag_train_test
            except: print(pre_text + 'En la variable flag_train_test hay que introducir tres '
            'cosas: el nombre de la variable con el flag, el valor de train y el valor de test.')

            data = X.copy()
            data[self.target_name] = y

            X_train = data[data[a] == b].drop(self.target_name, axis=1)
            y_train = data[data[a] == b][self.target_name].values

            X_test = data[data[a] == c].drop(self.target_name, axis=1)
            y_test = data[data[a] == c][self.target_name].values

        else:

            if self.stratify:
                if self.stratify_var == '':
                    X_train, X_test, y_train, y_test = train_test_split(X, y,
                    test_size=self.test_size, random_state=self.seed, stratify=y)
                else:
                    X_train, X_test, y_train, y_test = train_test_split(X, y,
                    test_size=self.test_size, random_state=self.seed, stratify=X[self.stratify_var])

            else:
                X_train, X_test, y_train, y_test = train_test_split(
                X, y, test_size=self.test_size, random_state=self.seed)

        self.index_train, self.index_test = X_train.index, X_test.index

        X_train, X_test = X_train.reset_index(drop=True), X_test.reset_index(drop=True)
        if isinstance(y, pd.Series): y_train, y_test = y_train.values, y_test.values

        self.y_train, self.y_test = y_train, y_test

        if self.flag_train_test == []:
            if self.stratify:
                if self.stratify_var == '':
                    print(pre_text + 'Particionado {}-{} estratificado en el target terminado.'\
                    .format(int(100*(1-self.test_size)), int(100*self.test_size)))
                    print(pre_text + '-' * N)
                else:
                    print(pre_text + 'Particionado {}-{} estratificado'
                    ' en la variable \'{}\' terminado.'.format(
                    int(100*(1-self.test_size)), int(100*self.test_size), self.stratify_var))
                    print(pre_text + '-' * N)
            else:
                print(pre_text + 'Particionado {}-{} terminado.'\
                .format(int(100*(1-self.test_size)), int(100*self.test_size)))
                print(pre_text + '-' * N)

        if self.features != []: variables = self.features

        else:
            if self.candidate_vars != []:
                variables = list(set(self.candidate_vars) - set(self.excluded_vars))
            else: variables = list(set(list(X.columns)) - set(self.excluded_vars))

        objetos = {}
        variables_no_agrupadas = []
        for variable in variables:

            try:
                x = X_train[variable].values
                frenken = auto_grouping(name=variable, max_groups=self.autogrp_max_groups,
                min_pct=self.autogrp_min_pct).fit(x, y_train)
                objetos[variable] = frenken

            except: variables_no_agrupadas.append(variable)

        print(pre_text + 'Autogrouping terminado. Máximo número de buckets = {}. '
        'Mínimo porcentaje por bucket = {}'.format(self.autogrp_max_groups, self.autogrp_min_pct))
        print(pre_text + '-' * N)

        if variables_no_agrupadas != []:

            print(pre_text + 'Variables no agrupadas: {}'.format(variables_no_agrupadas))
            print(pre_text + '-' * N)

            variables_booleanas = []
            for variable in variables_no_agrupadas:
                if True in X_train[variable].values and False in X_train[variable].values:
                    variables_booleanas.append(variable)

            if len(variables_booleanas) > 0:
                print(pre_text + 'Cuidado! Estas variables son booleanas, para usarlas '
                'transfórmalas a string antes: {}'.format(variables_booleanas))
                print(pre_text + '-' * N)

        tabla_ivs, contador = pd.DataFrame(columns=['variable', 'iv']), 0
        for variable in objetos:
            tabla_ivs.loc[contador] = variable, objetos[variable].iv
            contador += 1

        tabla_ivs = tabla_ivs.sort_values('iv', ascending=False)
        variables_filtroiv = tabla_ivs[tabla_ivs['iv'] >= 0.015]['variable']

        self.tabla_ivs = tabla_ivs

        variables_def = list(set(variables_filtroiv) - set(variables_no_agrupadas))
        self.final_breakpoints = compute_final_breakpoints(
        variables_def, objetos, self.user_breakpoints)
        
        info = compute_info(X_train, variables_def, self.final_breakpoints)
        df_train = adapt_data(X_train, y_train,
        variables_def, self.final_breakpoints, self.target_name)
        df_test = adapt_data(X_test, y_test,
        variables_def, self.final_breakpoints, self.target_name)

        features = self.features

        if self.selection_muestra_test: muestra_test = df_test
        else: muestra_test = None

        features = features_selection(
        df_train, self.features, variables_def, info, self.target_name,
        method=self.selection_method, metric=self.selection_metric,
        threshold=self.selection_threshold, max_iters=self.selection_max_iters,
        included_vars=self.included_vars, muestra_test=muestra_test, show=self.selection_show)

        df_train = df_train[features + [self.target_name]]
        df_test = df_test[features + [self.target_name]]

        scorecard, features_length = compute_scorecard(
        df_train, features, info, target_name=self.target_name)

        df_train_final, ks_train, gini_train = apply_scorecard(
        df_train, scorecard, info, metrics=['gini', 'ks'], target_name=self.target_name)
        print(pre_text + 'El modelo tiene un {:.2f}% de KS y un {:.2f}% de Gini en '
        'la muestra de desarrollo'.format(round(ks_train*100, 2), round(gini_train*100, 2)))
        print(pre_text + '-' * N)

        df_test_final, ks_test, gini_test = apply_scorecard(
        df_test, scorecard, info, metrics=['gini', 'ks'], target_name=self.target_name)
        print(pre_text + 'El modelo tiene un {:.2f}% de KS y un {:.2f}% de Gini en '
        'la muestra de validación'.format(round(ks_test*100, 2), round(gini_test*100, 2)))
        print(pre_text + '-' * N)

        self.variables_no_agrupadas = variables_no_agrupadas
        self.features = features
        self.scorecard = scorecard
        self.features_length = features_length
        self.ks_train = ks_train
        self.gini_train = gini_train
        self.ks_test = ks_test
        self.gini_test = gini_test
        
        if self.create_excel:
            
            try: self.create_sreadsheet()
            except: 
                print(pre_text + 'Por algún motivo no se ha podido '
                'generar el excel. ¿Tienes instalada la librería openpyxl')
                print(pre_text + '-' * N)

        if self.save_whole_tables: self.X_train, self.X_test = X_train, X_test
        else:
            self.X_train = X_train[self.id_columns + features]
            self.X_test = X_test[self.id_columns + features]

        for objeto in objetos: del objetos[objeto].x_final
        if self.save_all_autogroupings: self.objetos = objetos
        else: self.objetos = dict((k, objetos[k]) for k in features if k in objetos)

        self.pyspark_formula = compute_pyspark_formula(self)

        return self
    

    def transform(self, X, id_columns=[], target_name='',
    binary_prediction=True, metrics=[], print_log=True):

        if self.log_mode: pre_text = '123: '
        else: pre_text = ''

        if isinstance(X, pd.DataFrame):

            if target_name != '': X1 = X[id_columns + self.features + [target_name]].copy()
            else: X1 = X[id_columns + self.features].copy()

            X1_v2, info = X1.copy(), {}

            for feature in self.features:

                info[feature] = {}

                bp = self.final_breakpoints[feature]

                if not isinstance(bp, dict):
                    X1_v2[feature] = data_convert(
                    string_categories2(bp)).fit(X1[feature].values).x_final
                    info[feature]['breakpoints_num'] = breakpoints_to_num(bp)
                    info[feature]['group_names'] = compute_group_names(X1[feature].values.dtype, bp)

                else:
                    X1_v2[feature] = remapeo_missing(
                    data_convert(string_categories2(bp)).fit(X1[feature].values).x_final, bp)
                    info[feature]['breakpoints_num'] = breakpoints_to_num(bp['breakpoints'])
                    info[feature]['group_names'] = compute_group_names(
                    X1[feature].values.dtype, bp['breakpoints'], bp['missing_group'])

            salida = apply_scorecard(X1_v2, self.scorecard, info, 
            binary_prediction=binary_prediction, metrics=metrics,
            target_name=target_name, print_log=print_log)

            if metrics == []: X2 = salida
            else: X2 = salida[0]

            venga = 0
            for i in X2.columns:
                if 'scr_' in i:
                    break
                venga += 1

            for i in X2.columns[venga:]: X1[i] = X2[i]

            if metrics == []: return X1
            else: return X1, salida[1], salida[2]

        else:
            
            import pyspark.sql.functions as sf
            from pyspark.sql.types import DoubleType
            from pyspark.ml.evaluation import BinaryClassificationEvaluator

            if target_name != '':
                X1 = X.select(id_columns + self.features + [target_name])\
                .withColumn('scorecardpoints', sf.lit(0.0).cast(DoubleType()))
            else:
                X1 = X.select(id_columns + self.features)\
                .withColumn('scorecardpoints', sf.lit(0.0).cast(DoubleType()))

            for i in range(len(self.pyspark_formula)):
                X1 = X1.withColumn('scr_{}'.format(self.features[i]),
                sf.expr(self.pyspark_formula[i]))\
                .withColumn('scorecardpoints',
                sf.col('scorecardpoints') + sf.col('scr_{}'.format(self.features[i])))

            if binary_prediction:
                X1 = X1.withColumn('prediction',
                sf.when(sf.col('scorecardpoints') >= 500, 0).otherwise(1))

            columnas = list(X1.columns).copy()
            columnas.remove('scorecardpoints')
            if binary_prediction: columnas.remove('prediction')
            columnas += ['scorecardpoints']
            if binary_prediction: columnas += ['prediction']
            X1 = X1.select(columnas)

            if metrics == []: return X1

            else:

                if metrics not in (['ks'], ['gini'], ['ks', 'gini'], ['gini', 'ks']):
                    raise ValueError("Valor erroneo para 'metrics'. Los valores "
                    "váidos son: ['ks'], ['gini'], ['ks', 'gini'], ['gini', 'ks']")

                if target_name == '':
                    raise ValueError("Si el parámetro 'metrics' viene relleno entonces "
                    "debe especificarse el nombre de la variable objetivo en 'target_name'")

                if 'ks' in metrics:
                    ks = calcula_ks_pyspark(X1, target_name, 'scorecardpoints')[1]

                if 'gini' in metrics:
                    evaluator = BinaryClassificationEvaluator(rawPredictionCol='scorecardpoints',
                    labelCol=target_name, metricName='areaUnderROC')
                    auroc = evaluator.evaluate(X1)
                    gini = 1 - 2 * auroc

                if 'ks' in metrics and 'gini' not in metrics:
                    if print_log:
                        print(pre_text + 'El modelo tiene un {:.2f}% de KS '
                        'en esta muestra'.format(round(ks*100, 2)))
                    return X1, ks

                if 'ks' not in metrics and 'gini' in metrics:
                    if print_log:
                        print(pre_text + 'El modelo tiene un {:.2f}% de Gini '
                        'en esta muestra'.format(round(gini*100, 2), ))
                    return X1, gini

                if 'ks' in metrics and 'gini' in metrics:
                    if print_log:
                        print(pre_text + 'El modelo tiene un {:.2f}% de KS y un {:.2f}% de Gini '
                        'en esta muestra'.format(round(ks*100, 2), round(gini*100, 2)))
                    return X1, ks, gini
                
                
    def create_sreadsheet(self):
        
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        scorecard = self.scorecard.copy()
        scorecard = scorecard.drop('Raw score', axis=1)
        
        abc = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
        'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']

        wb = openpyxl.Workbook()
        ws0 = wb['Sheet']
        rows = dataframe_to_rows(scorecard, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                try: ws0.cell(row=r_idx, column=c_idx, value=value)
                except: ws0.cell(row=r_idx, column=c_idx, value=str(value))

        ws0.insert_cols(2)
        ws0.insert_cols(4)
        ws0.insert_cols(12)

        ws0.merge_cells('A1:B1')
        ws0.merge_cells('C1:D1')

        altura = len(scorecard)

        for letra in ['F', 'I']:
            for row in ws0['{}2:{}{}'.format(letra, letra, altura+1)]:
                for cell in row:
                    cell.number_format = '0.00%'

        for letra in ['J', 'K', 'L']:
            for row in ws0['{}2:{}{}'.format(letra, letra, altura+1)]:
                for cell in row:
                    cell.number_format = '0.0000'

        for i in range(2, altura+2):
            ws0.merge_cells('C{}:D{}'.format(i, i))

        for row in ws0['A1:M1']:
            for cell in row:
                cell_style(cell, bold=True, hor_alignment='center', ver_alignment='center', 
                all_borders=True, font_color='ffffff', background_color='ff0000')

        for row in ws0['A2:M{}'.format(altura+1)]:
            for cell in row:
                cell_style(cell, hor_alignment='center', ver_alignment='center', all_borders=True, wrap_text=True)

        ws0['K1'].value = 'IV aux'
        ws0['L1'].value = 'IV'

        contador = 2
        for i in self.features_length:
            new_contador = contador+i
            ws0.merge_cells('A{}:B{}'.format(contador, new_contador-1))
            ws0['L{}'.format(contador)] = '=SUM(K{}:K{})'.format(contador, new_contador-1)
            ws0.merge_cells('L{}:L{}'.format(contador, new_contador-1))
            contador = new_contador

        for letra in abc: ws0.column_dimensions[letra].width = 12.89
        ws0.sheet_view.showGridLines = False
        ws0.column_dimensions['N'].width = 8
        ws0.column_dimensions['K'].hidden= True
        ws0.sheet_view.zoomScale = 85

        ws0['O3'].value = 'KS'
        ws0['O4'].value = 'GINI'
        ws0['P2'].value = 'Train'
        ws0['Q2'].value = 'Test'
        ws0['P3'].value = self.ks_train
        ws0['Q3'].value = self.ks_test
        ws0['P4'].value = self.gini_train
        ws0['Q4'].value = self.gini_test

        for celda in ['P2', 'Q2', 'O3', 'O4']:
            cell_style(ws0[celda], bold=True, hor_alignment='center', ver_alignment='center', 
            all_borders=True, font_color='ffffff', background_color='ff0000')

        for celda in ['P3', 'Q3', 'P4', 'Q4']:
            cell_style(ws0[celda], hor_alignment='center', ver_alignment='center', all_borders=True, wrap_text=True)
            ws0[celda].number_format = '0.00%'
            
        self.excel = wb
        

    def save_excel(self, ruta, color='blue'):
        
        import openpyxl
        from openpyxl.styles import PatternFill
        
        if color == 'green': color = 'CCFFCC'
        if color == 'light_blue': color = 'CCFFFF'
        if color == 'blue': color = 'CCECFF'
        if color == 'pink': color = 'FFCCFF'
        if color == 'red': color = 'FFCCCC'
        if color == 'yellow': color = 'FFFFCC'
        if color == 'purple': color = 'CCCCFE'
        if color == 'orange': color = 'FFCC99'
        
        wb = self.excel
        ws0 = wb['Sheet']
        
        contador, moneda = 2, 0
        for i in self.features_length:
            new_contador = contador+i
            if moneda%2 == 0:
                for row in ws0['A{}:M{}'.format(contador, new_contador-1)]:
                    for cell in row:
                        cell.fill = PatternFill(fill_type='solid', fgColor=color)
            contador = new_contador
            moneda += 1
            
        wb.save(ruta)
        
    
class auto_grouping:


    def __init__(self, name, max_groups=5, min_pct=0.05, log_mode=False):

        self.name = name
        self.max_groups = max_groups
        self.min_pct = min_pct
        self.log_mode = log_mode
        

    def fit(self, x, y):
        
        N = 150
        if self.log_mode: pre_text = '123: '
        else: pre_text = ''

        dtype = x.dtype
        self.dtype = dtype

        if dtype != 'O': categories = {}
        else:
            categories = string_categories1(x, y)
            if pd.Series(x).isna().sum() > 0:
                for i in categories:
                    if not isinstance(i, str):
                        categories['Missing'] = categories.pop(i)
                categories = dict(sorted(categories.items(), key=lambda item: item[1]))

        self.categories = categories
        frenken = data_convert(categories).fit(x)
        x_converted = frenken.x_converted
        self.x_final = frenken.x_final

        if dtype != 'O' and np.isnan(x_converted).sum() > 0:
            aux = ~ np.isnan(x_converted)
            x_nm, y_nm = x_converted[aux], y[aux]

        else: x_nm, y_nm = x_converted, y
        self.compute_groups(x_nm, y_nm)

        if dtype != 'O' and np.isnan(x).sum() > 0:
            
            self.breakpoints_num = np.array([-12345670] + list(self.breakpoints_num))
                
            x_groups = np.digitize(self.x_final, self.breakpoints_num)
            ngroups = len(self.breakpoints_num) + 1
            g = np.empty(ngroups).astype(np.int64)
            b = np.empty(ngroups).astype(np.int64)
            for i in range(ngroups):
                g[i] = np.sum([(y == 0) & (x_groups == i)])
                b[i] = np.sum([(y == 1) & (x_groups == i)])
            
            if any(b == 0):
                print(pre_text + 'La variable {} no se ha podido agrupar porque '
                'en los missings no hay ni un solo malo'.format(self.name))
                print('-' * N)
                raise ValueError('Errorcito')
            
        if dtype == 'O':
            self.breakpoints = breakpoints_to_str(self.breakpoints_num, categories)
        else: self.breakpoints = self.breakpoints_num

        self.iv = compute_iv(self.x_final, y, self.breakpoints_num)

        group_names = compute_group_names(dtype, self.breakpoints, 0)
        self.table = compute_table(self.x_final, y, self.breakpoints_num, group_names)

        return self


    def compute_groups(self, x, y):

        tree = DecisionTreeClassifier(**{'min_samples_leaf': self.min_pct,
        'max_leaf_nodes': self.max_groups}).fit(x.reshape(-1, 1), y)
        aux = np.unique(tree.tree_.threshold)
        breakpoints_num = aux[aux != _tree.TREE_UNDEFINED]

        x_groups = np.digitize(x, breakpoints_num)

        ngroups = len(breakpoints_num) + 1
        g = np.empty(ngroups).astype(np.int64)
        b = np.empty(ngroups).astype(np.int64)

        for i in range(ngroups):

            g[i] = np.sum([(y == 0) & (x_groups == i)])
            b[i] = np.sum([(y == 1) & (x_groups == i)])

        error = (g == 0) | (b == 0)

        while np.any(error):

            m_bk = np.concatenate(
            [error[:-2], [error[-2] | error[-1]]])

            breakpoints_num = breakpoints_num[~m_bk]
            x_groups = np.digitize(x, breakpoints_num)

            ngroups = len(breakpoints_num) + 1
            g = np.empty(ngroups).astype(np.int64)
            b = np.empty(ngroups).astype(np.int64)

            for i in range(ngroups):
                g[i] = np.sum([(y == 0) & (x_groups == i)])
                b[i] = np.sum([(y == 1) & (x_groups == i)])

            error = (g == 0) | (b == 0)

        self.breakpoints_num = breakpoints_num


class data_convert:


    def __init__(self, categories={}, missing_value=-12345678):

        self.categories = categories
        self.missing_value = missing_value


    def fit(self, x):

        dtype = x.dtype
        self.x_original = x

        if dtype == 'O':

            if self.categories == {}:
                raise ValueError('En una variable de tipo texto es '
                'necesario especificar el diccionario de categorias')

            if pd.Series(x).isna().sum() > 0:
                x = pd.Series(x).replace(np.nan, 'Missing').values

            self.x_initial = x
            x = string_to_num(x, self.categories)
            self.x_converted = x

        else:
            self.x_initial = x
            self.x_converted = x

        if dtype != 'O' and np.isnan(x).sum() > 0:
            self.x_final = np.nan_to_num(x, nan=self.missing_value)
        else: self.x_final = self.x_converted

        return self


def adapt_data(X, y, variables, breakpoints, target_name='target'):
    
    df = pd.DataFrame()
    for variable in variables:
        
            bp = breakpoints[variable]
            
            if not isinstance(bp, dict):
                df[variable] = data_convert(
                string_categories2(bp)).fit(X[variable].values).x_final
                
            else:
                df[variable] = remapeo_missing(
                data_convert(string_categories2(bp)).fit(X[variable].values).x_final, bp)
        
    df[target_name] = y

    return df

